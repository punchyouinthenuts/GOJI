import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

# Validate job number (five-digit number)
def validate_job_number(job_number):
    return re.match(r'^\d{5}$', job_number) is not None

# Validate month (three-letter abbreviation)
def validate_month(month):
    months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
    return month.upper() in months

# Get domestic counts from CSV
def get_domestic_counts(file_path):
    df = pd.read_csv(file_path)
    df['State'] = df['City State ZIP Code'].str.split().str[-2]
    total_pieces = len(df)
    ny_pieces = df[df['State'].str.upper() == 'NY'].shape[0]
    return total_pieces, ny_pieces

# Get international counts from CSV
def get_international_counts(file_path):
    df = pd.read_csv(file_path)
    can_pieces = df[df['Country'].str.upper() == 'CANADA'].shape[0]
    nz_pieces = df[df['Country'].str.upper() == 'NEW ZEALAND'].shape[0]
    return can_pieces, nz_pieces

def main():
    # Input and validate job number
    job_number = input("ENTER JOB NUMBER: ")
    while not validate_job_number(job_number):
        print("Invalid job number. Please enter a five-digit number (e.g., 12345).")
        job_number = input("ENTER JOB NUMBER: ")

    # Input and validate month
    month = input("ENTER MONTH: ").upper()
    while not validate_month(month):
        print("Invalid month. Please enter a three-letter month abbreviation (e.g., MAR).")
        month = input("ENTER MONTH: ").upper()

    # Input file paths
    domestic_path = input("INPUT THE DOMESTIC LIST: ").strip('"')
    while not os.path.exists(domestic_path):
        print("File not found. Please enter a valid file path.")
        domestic_path = input("INPUT THE DOMESTIC LIST: ").strip('"')

    international_path = input("INPUT THE INTERNATIONAL LIST: ").strip('"')
    while not os.path.exists(international_path):
        print("File not found. Please enter a valid file path.")
        international_path = input("INPUT THE INTERNATIONAL LIST: ").strip('"')

    # Input total domestic postage
    try:
        total_domestic_postage = float(input("INPUT TOTAL DOMESTIC POSTAGE: ").replace('$', '').replace(',', ''))
        total_domestic_postage = round(total_domestic_postage, 2)
    except ValueError:
        print("Invalid postage amount. Please enter a numeric value.")
        return

    # Input per-piece costs
    try:
        can_per_piece = float(input("INPUT COST OF CAN POSTAGE PER PIECE: "))
        nz_per_piece = float(input("INPUT COST OF NZ POSTAGE PER PIECE: "))
    except ValueError:
        print("Invalid per-piece cost. Please enter a numeric value.")
        return

    # Process data
    total_pieces, ny_pieces = get_domestic_counts(domestic_path)
    us_pieces = total_pieces - ny_pieces
    average_per_piece = round(total_domestic_postage / total_pieces, 2) if total_pieces > 0 else 0
    us_total_postage = round(us_pieces * average_per_piece, 2)
    ny_total_postage = round(ny_pieces * average_per_piece, 2)
    can_pieces, nz_pieces = get_international_counts(international_path)
    can_total_postage = round(can_pieces * can_per_piece, 2)
    nz_total_postage = round(nz_pieces * nz_per_piece, 2)
    combined_total = round(total_domestic_postage + can_total_postage + nz_total_postage, 2)
    international_total = round(can_total_postage + nz_total_postage, 2)

    # Data structure for Excel
    data = [
        [job_number, f"AILI SPOTLIGHT {month}", total_domestic_postage, total_pieces, average_per_piece, "STD", "FLT", "1165"],
        ["", "US", us_total_postage, us_pieces, average_per_piece, "STD", "FLT", "1165"],
        ["", "NY", ny_total_postage, ny_pieces, average_per_piece, "STD", "FLT", "1165"],
        ["", "CAN", can_total_postage, can_pieces, can_per_piece, "FC", "FLT", "METERED"],
        ["", "NZ", nz_total_postage, nz_pieces, nz_per_piece, "FC", "FLT", "METERED"],
        ["", "COMBINED TOTAL", combined_total, "", "", "", "", ""],
        ["", "INTERNATIONAL TOTAL", international_total, "", "", "", "", ""]
    ]

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active

    # Define border style
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Write data with formatting
    for row_idx, row in enumerate(data, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if value != "":
                cell.value = value
                if col_idx == 3:  # Column C: postage amounts
                    cell.number_format = '$#,##0.00'
                elif col_idx == 5 and row_idx <= 5:  # Column E: averages for rows 1-5
                    cell.number_format = '$#,##0.00'
                elif col_idx == 4:  # Column D: piece counts
                    cell.number_format = '0'
                cell.border = border

    # Adjust column widths
    column_widths = []
    for row in data:
        for i, value in enumerate(row):
            if value != "":
                length = len(str(value))
                if i < len(column_widths):
                    column_widths[i] = max(column_widths[i], length)
                else:
                    column_widths.append(length)
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width + 2

    # Save and open file
    output_path = r"C:\Users\JCox\Downloads\postage_report.xlsx"
    wb.save(output_path)
    os.startfile(output_path)
    print(f"Excel file generated and saved at: {output_path}")

if __name__ == "__main__":
    main()
